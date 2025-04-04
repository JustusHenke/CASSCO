import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import os
from matplotlib.colors import LinearSegmentedColormap
import matplotlib.gridspec as gridspec
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule

# Create simulations directory if it doesn't exist
if not os.path.exists('simulations'):
    os.makedirs('simulations')

# Model parameters class with scenario-specific settings
class ModelParams:
    def __init__(self, scenario='gen_ai'):
        # Common parameters for both scenarios
        self.P_altruistic = 0.3  # Private benefit for altruistic strategies
        self.P_egoistic = 0.7    # Private benefit for egoistic strategies
        
        # Set the scenario
        self.scenario = scenario
        
        # Apply scenario-specific parameters
        self.set_scenario_params(scenario)
    
    def set_scenario_params(self, scenario):
        """Set all parameters based on scenario"""
        if scenario == 'gen_ai':
            # Generative AI scenario parameters
            self.M_O = 0.5       # Max impact for organization
            self.M_R = 0.5       # Max impact for researcher
            self.alpha_O = 0.5   # Weight for organization (impact vs private benefit)
            self.alpha_R = 0.5   # Weight for researcher
            self.X_init = 0.5    # Initial exigence
            self.beta = 0.4      # Learning rate
            self.theta = 0.3     # Threshold impact
        elif scenario == 'citizen_science':
            # Citizen Science scenario parameters
            self.M_O = 0.4       # Max impact for organization
            self.M_R = 0.6       # Max impact for researcher
            self.alpha_O = 0.6   # Weight for organization (impact vs private benefit)
            self.alpha_R = 0.7   # Weight for researcher
            self.X_init = 0.2    # Initial exigence
            self.beta = 0.3      # Learning rate
            self.theta = 0.5     # Threshold impact
        else:
            raise ValueError(f"Unknown scenario: {scenario}")

# Define the possible strategies
def get_strategies():
    roles = ['a', 'e']  # a:altruistic, e:egoistic
    modes = ['k', 'd', 'p']  # k:knowledge, d:dialogue, p:participation
    return [(r, m) for r in roles for m in modes]

# Get impact values based on scenario
def get_impact_values(scenario):
    if scenario == 'gen_ai':
        # Impact assessments for the generative AI scenario (from Table 5 in the paper)
        impact_O = {
            ('a', 'k'): 0.7,  # Detailed explanations of AI methods
            ('a', 'd'): 0.5,  # Expert panels on AI validation
            ('a', 'p'): 0.4,  # Public AI validation workshops
            ('e', 'k'): 0.6,  # Presentation of AI validation protocols
            ('e', 'd'): 0.4,  # AI ethics committees
            ('e', 'p'): 0.3   # AI hackathons for validation
        }
        
        impact_R = {
            ('a', 'k'): 0.9,  # Scientific publications on AI validation methods
            ('a', 'd'): 0.7,  # Peer-review discussions on AI methods
            ('a', 'p'): 0.5,  # Open-source validation tools
            ('e', 'k'): 0.8,  # Longitudinal studies on AI's effect
            ('e', 'd'): 0.4,  # AI-focused conference contributions
            ('e', 'p'): 0.3   # Crowdsourcing of AI training
        }
    elif scenario == 'citizen_science':
        # Impact assessments for the citizen science scenario (from Table 4 in the paper)
        impact_O = {
            ('a', 'k'): 0.5,  # Providing resources and expert knowledge
            ('a', 'd'): 0.7,  # Organizing discussion forums and workshops
            ('a', 'p'): 0.9,  # Providing infrastructure for citizen projects
            ('e', 'k'): 0.3,  # Dissemination of research results
            ('e', 'd'): 0.5,  # Networking events with stakeholders
            ('e', 'p'): 0.6   # Large-scale citizen science campaigns
        }
        
        impact_R = {
            ('a', 'k'): 0.5,  # Creation of educational materials
            ('a', 'd'): 0.7,  # Conducting citizen dialogues
            ('a', 'p'): 0.9,  # Joint data collection and analysis with citizens
            ('e', 'k'): 0.3,  # Writing academic articles
            ('e', 'd'): 0.5,  # Presentations at conferences
            ('e', 'p'): 0.6   # Leading citizen science projects
        }
    else:
        raise ValueError(f"Unknown scenario: {scenario}")
    
    return impact_O, impact_R

# Calculate Bayesian likelihood for belief update
def calculate_likelihood(strategy, observed_impact, normalized_impact, strategies):
    """
    Calculate likelihood P(I*|s_j) based on how well the strategy contributes to observed impact
    """
    # Calculate how much this strategy contributes to the total possible impact
    strategy_impact = normalized_impact[strategy]
    total_impact = sum(normalized_impact.values())
    
    # Calculate relative contribution of this strategy
    contribution = strategy_impact / total_impact if total_impact > 0 else 0
    
    # Calculate likelihood based on the contribution and observed impact
    # Higher contribution → higher likelihood
    likelihood = contribution * observed_impact
    
    # Ensure likelihood is positive and normalized
    likelihood = max(0.001, likelihood)  # Avoid zero likelihood
    
    return likelihood

# Calculate one step of the game theoretic model
def calculate_game_step(params, beliefs_O=None, beliefs_R=None, t=0, export_excel=False, prev_beliefs_O=None, prev_beliefs_R=None):
    """
    Calculates one step of the game theoretic model following the paper's formulation
    """
    strategies = get_strategies()
    
    # Get impact values for the selected scenario
    impact_O, impact_R = get_impact_values(params.scenario)
    
    # Normalize impacts according to equation 4
    normalized_impact_O = {k: v * params.M_O for k, v in impact_O.items()}
    normalized_impact_R = {k: v * params.M_R for k, v in impact_R.items()}
    
    # Initialize beliefs based on equation 5 if not provided
    if beliefs_O is None or beliefs_R is None:
        sum_impact_O = sum(normalized_impact_O.values())
        sum_impact_R = sum(normalized_impact_R.values())
        
        beliefs_O = {s: normalized_impact_R[s]/sum_impact_R for s in strategies}
        beliefs_R = {s: normalized_impact_O[s]/sum_impact_O for s in strategies}
    
    # Calculate expected impacts according to equation 6
    expected_impacts_O = {}
    expected_impacts_R = {}
    
    for s_O in strategies:
        # Expected impact of O's strategy
        own_impact = normalized_impact_O[s_O]
        other_expected_impact = sum(beliefs_O[s_R] * normalized_impact_R[s_R] for s_R in strategies)
        expected_impacts_O[s_O] = own_impact + other_expected_impact
    
    for s_R in strategies:
        # Expected impact of R's strategy
        own_impact = normalized_impact_R[s_R]
        other_expected_impact = sum(beliefs_R[s_O] * normalized_impact_O[s_O] for s_O in strategies)
        expected_impacts_R[s_R] = own_impact + other_expected_impact
    
    # Calculate expected utilities according to equation 7
    expected_utilities_O = {}
    expected_utilities_R = {}
    
    for s_O in strategies:
        private_benefit = params.P_egoistic if s_O[0] == 'e' else params.P_altruistic
        expected_utilities_O[s_O] = params.alpha_O * expected_impacts_O[s_O] + (1 - params.alpha_O) * private_benefit
    
    for s_R in strategies:
        private_benefit = params.P_egoistic if s_R[0] == 'e' else params.P_altruistic
        expected_utilities_R[s_R] = params.alpha_R * expected_impacts_R[s_R] + (1 - params.alpha_R) * private_benefit
    
    # Choose optimal strategies according to equation 8
    s_O_opt = max(expected_utilities_O, key=expected_utilities_O.get)
    s_R_opt = max(expected_utilities_R, key=expected_utilities_R.get)
    
    # Calculate observed impact based on optimal strategies with random variation
    import random
    
    # Apply randomization similar to Excel formula: MIN(MAX(base_value + RAND()*0.2-0.1, 0), 1)
    observed_impact_O = min(max(normalized_impact_O[s_O_opt] + random.random() * 0.2 - 0.1, 0), 1)
    observed_impact_R = min(max(normalized_impact_R[s_R_opt] + random.random() * 0.2 - 0.1, 0), 1)
    
    # Ensure impacts don't exceed maximum possible values
    observed_impact_O = min(observed_impact_O, params.M_O)
    observed_impact_R = min(observed_impact_R, params.M_R)
    
    # Calculate total observed impact
    observed_total_impact = observed_impact_O + observed_impact_R
    
    # Update beliefs according to equation 9 (Bayesian update)
    # Calculate likelihoods for each strategy
    likelihood_O = {}
    likelihood_R = {}
    
    for s in strategies:
        # Calculate likelihood P(I*|s) for Organization's beliefs about Researcher strategies
        likelihood_O[s] = calculate_likelihood(s, observed_total_impact, normalized_impact_R, strategies)
        
        # Calculate likelihood P(I*|s) for Researcher's beliefs about Organization strategies
        likelihood_R[s] = calculate_likelihood(s, observed_total_impact, normalized_impact_O, strategies)
    
    # Calculate unnormalized posterior beliefs
    unnormalized_posterior_O = {s: likelihood_O[s] * beliefs_O[s] for s in strategies}
    unnormalized_posterior_R = {s: likelihood_R[s] * beliefs_R[s] for s in strategies}
    
    # Calculate normalization factors
    normalization_factor_O = sum(unnormalized_posterior_O.values())
    normalization_factor_R = sum(unnormalized_posterior_R.values())
    
    # Calculate updated beliefs (normalized posterior)
    updated_beliefs_O = {s: v/normalization_factor_O for s, v in unnormalized_posterior_O.items()}
    updated_beliefs_R = {s: v/normalization_factor_R for s, v in unnormalized_posterior_R.items()}
    
    
    # Recalculate expected impacts with updated beliefs
    recalculated_impacts_O = {}
    recalculated_impacts_R = {}
    
    for s_O in strategies:
        # Expected impact of O's strategy with updated beliefs
        own_impact = normalized_impact_O[s_O]
        other_expected_impact = sum(updated_beliefs_O[s_R] * normalized_impact_R[s_R] for s_R in strategies)
        recalculated_impacts_O[s_O] = own_impact + other_expected_impact
    
    for s_R in strategies:
        # Expected impact of R's strategy with updated beliefs
        own_impact = normalized_impact_R[s_R]
        other_expected_impact = sum(updated_beliefs_R[s_O] * normalized_impact_O[s_O] for s_O in strategies)
        recalculated_impacts_R[s_R] = own_impact + other_expected_impact
    
    # Recalculate expected utilities with updated beliefs
    recalculated_utilities_O = {}
    recalculated_utilities_R = {}
    
    for s_O in strategies:
        private_benefit = params.P_egoistic if s_O[0] == 'e' else params.P_altruistic
        recalculated_utilities_O[s_O] = params.alpha_O * recalculated_impacts_O[s_O] + (1 - params.alpha_O) * private_benefit
    
    for s_R in strategies:
        private_benefit = params.P_egoistic if s_R[0] == 'e' else params.P_altruistic
        recalculated_utilities_R[s_R] = params.alpha_R * recalculated_impacts_R[s_R] + (1 - params.alpha_R) * private_benefit
    
    
    # Update exigence according to equation 10
    X_next = params.X_init + params.beta * (params.theta - observed_total_impact) * params.X_init * (1 - params.X_init)
    exigence_change = X_next - params.X_init
    
    # Export calculations to Excel if requested
    if export_excel:
        scenario_name = 'GenerativeAI' if params.scenario == 'gen_ai' else 'CitizenScience'
        excel_filename = f'simulations/{scenario_name}_Calculations_Step{t}.xlsx'
        
        # Create a workbook and sheets
        workbook = openpyxl.Workbook()
        
        # Model Parameters Sheet (NEW)
        params_sheet = workbook.active
        params_sheet.title = "Model Parameters"
        params_sheet['A1'] = "Model Parameters and Assumptions"
        params_sheet['A1'].font = Font(bold=True, size=14)
        
        # Add model parameters
        row = 3
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        params_sheet['A3'] = "Parameter"
        params_sheet['B3'] = "Value"
        params_sheet['C3'] = "Description"
        
        params_sheet['A3'].font = Font(bold=True)
        params_sheet['B3'].font = Font(bold=True)
        params_sheet['C3'].font = Font(bold=True)
        
        # Apply borders to headers
        params_sheet['A3'].border = border
        params_sheet['B3'].border = border
        params_sheet['C3'].border = border
        
        # Add all parameters
        parameter_info = [
            ('Scenario', params.scenario, 'Generative AI or Citizen Science scenario'),
            ('M_O', params.M_O, 'Maximum impact for organization'),
            ('M_R', params.M_R, 'Maximum impact for researcher'),
            ('alpha_O', params.alpha_O, 'Weight between impact and private benefit for organization'),
            ('alpha_R', params.alpha_R, 'Weight between impact and private benefit for researcher'),
            ('P_altruistic', params.P_altruistic, 'Private benefit for altruistic strategies'),
            ('P_egoistic', params.P_egoistic, 'Private benefit for egoistic strategies'),
            ('X_init', params.X_init, 'Initial exigence value'),
            ('beta', params.beta, 'Learning rate for exigence updates'),
            ('theta', params.theta, 'Threshold impact value')
        ]
        
        for i, (param, value, desc) in enumerate(parameter_info):
            row = i + 4
            params_sheet[f'A{row}'] = param
            params_sheet[f'B{row}'] = value
            params_sheet[f'C{row}'] = desc
            
            # Apply borders
            params_sheet[f'A{row}'].border = border
            params_sheet[f'B{row}'].border = border
            params_sheet[f'C{row}'].border = border
        
        # Add impact values section
        row += 2
        params_sheet[f'A{row}'] = "Impact Values"
        params_sheet[f'A{row}'].font = Font(bold=True, size=12)
        
        # Add Organization impact values
        row += 2
        params_sheet[f'A{row}'] = "Organization Impact Values"
        params_sheet[f'A{row}'].font = Font(bold=True)
        
        # Headers for impact values
        row += 1
        params_sheet[f'A{row}'] = "Strategy"
        params_sheet[f'B{row}'] = "Impact"
        params_sheet[f'C{row}'] = "Normalized"
        
        params_sheet[f'A{row}'].font = Font(bold=True)
        params_sheet[f'B{row}'].font = Font(bold=True)
        params_sheet[f'C{row}'].font = Font(bold=True)
        
        # Add each impact value
        for i, s in enumerate(strategies):
            row += 1
            params_sheet[f'A{row}'] = f"{s[0]},{s[1]}"
            params_sheet[f'B{row}'] = impact_O[s]
            params_sheet[f'C{row}'] = normalized_impact_O[s]
        
        # Add Researcher impact values
        row += 2
        params_sheet[f'A{row}'] = "Researcher Impact Values"
        params_sheet[f'A{row}'].font = Font(bold=True)
        
        # Headers for impact values
        row += 1
        params_sheet[f'A{row}'] = "Strategy"
        params_sheet[f'B{row}'] = "Impact"
        params_sheet[f'C{row}'] = "Normalized"
        
        params_sheet[f'A{row}'].font = Font(bold=True)
        params_sheet[f'B{row}'].font = Font(bold=True)
        params_sheet[f'C{row}'].font = Font(bold=True)
        
        # Add each impact value
        for i, s in enumerate(strategies):
            row += 1
            params_sheet[f'A{row}'] = f"{s[0]},{s[1]}"
            params_sheet[f'B{row}'] = impact_R[s]
            params_sheet[f'C{row}'] = normalized_impact_R[s]
            
        # Add beliefs section
        row += 2
        params_sheet[f'A{row}'] = "Current Beliefs"
        params_sheet[f'A{row}'].font = Font(bold=True, size=12)
        
        # If this is step 0, show "Initial Beliefs"
        # If not, show "Updated Beliefs from Previous Step"
        if t == 0:
            row += 2
            params_sheet[f'A{row}'] = "Initial Beliefs"
            params_sheet[f'A{row}'].font = Font(bold=True)
        else:
            row += 2
            params_sheet[f'A{row}'] = "Updated Beliefs from Previous Step"
            params_sheet[f'A{row}'].font = Font(bold=True)
            
            # If we have previous beliefs, show them for comparison
            if prev_beliefs_O is not None and prev_beliefs_R is not None:
                # Organization's beliefs about Researcher
                row += 2
                params_sheet[f'A{row}'] = "Organization's beliefs about Researcher strategies"
                params_sheet[f'A{row}'].font = Font(bold=True)
                
                row += 1
                params_sheet[f'A{row}'] = "Strategy"
                for i, s in enumerate(strategies):
                    params_sheet.cell(row=row, column=i+2).value = f"{s[0]},{s[1]}"
                
                row += 1
                for i, s in enumerate(strategies):
                    params_sheet.cell(row=row, column=i+2).value = prev_beliefs_O[s]
                
                # Researcher's beliefs about Organization
                row += 2
                params_sheet[f'A{row}'] = "Researcher's beliefs about Organization strategies"
                params_sheet[f'A{row}'].font = Font(bold=True)
                
                row += 1
                params_sheet[f'A{row}'] = "Strategy"
                for i, s in enumerate(strategies):
                    params_sheet.cell(row=row, column=i+2).value = f"{s[0]},{s[1]}"
                
                row += 1
                for i, s in enumerate(strategies):
                    params_sheet.cell(row=row, column=i+2).value = prev_beliefs_R[s]
        
        # Initial Beliefs Sheet
        beliefs_sheet = workbook.create_sheet("Initial Beliefs")
        beliefs_sheet['A1'] = "Initial Beliefs"
        beliefs_sheet['A1'].font = Font(bold=True, size=14)
        
        # Organization beliefs
        beliefs_sheet['A3'] = "Organization's beliefs about Researcher strategies"
        beliefs_sheet['A3'].font = Font(bold=True)
        beliefs_sheet['A4'] = "Strategy"
        for i, s in enumerate(strategies):
            beliefs_sheet.cell(row=4, column=i+2).value = f"{s[0]},{s[1]}"
            beliefs_sheet.cell(row=5, column=i+2).value = beliefs_O[s]
        
        # Researcher beliefs
        beliefs_sheet['A7'] = "Researcher's beliefs about Organization strategies"
        beliefs_sheet['A7'].font = Font(bold=True)
        beliefs_sheet['A8'] = "Strategy"
        for i, s in enumerate(strategies):
            beliefs_sheet.cell(row=8, column=i+2).value = f"{s[0]},{s[1]}"
            beliefs_sheet.cell(row=9, column=i+2).value = beliefs_R[s]
        
        # Expected Impacts Sheet
        impacts_sheet = workbook.create_sheet("Expected Impacts")
        impacts_sheet['A1'] = "Expected Impacts"
        impacts_sheet['A1'].font = Font(bold=True, size=14)
        
        # Organization expected impacts
        impacts_sheet['A3'] = "Organization's expected impacts"
        impacts_sheet['A3'].font = Font(bold=True)
        impacts_sheet['A4'] = "Strategy"
        for i, s in enumerate(strategies):
            impacts_sheet.cell(row=4, column=i+2).value = f"{s[0]},{s[1]}"
            impacts_sheet.cell(row=5, column=i+2).value = expected_impacts_O[s]
        
        # Researcher expected impacts
        impacts_sheet['A7'] = "Researcher's expected impacts"
        impacts_sheet['A7'].font = Font(bold=True)
        impacts_sheet['A8'] = "Strategy"
        for i, s in enumerate(strategies):
            impacts_sheet.cell(row=8, column=i+2).value = f"{s[0]},{s[1]}"
            impacts_sheet.cell(row=9, column=i+2).value = expected_impacts_R[s]
        
        # Expected Utilities Sheet
        utilities_sheet = workbook.create_sheet("Expected Utilities")
        utilities_sheet['A1'] = "Expected Utilities"
        utilities_sheet['A1'].font = Font(bold=True, size=14)
        
        # Organization expected utilities
        utilities_sheet['A3'] = "Organization's expected utilities"
        utilities_sheet['A3'].font = Font(bold=True)
        utilities_sheet['A4'] = "Strategy"
        for i, s in enumerate(strategies):
            utilities_sheet.cell(row=4, column=i+2).value = f"{s[0]},{s[1]}"
            utilities_sheet.cell(row=5, column=i+2).value = expected_utilities_O[s]
            # Highlight optimal strategy
            if s == s_O_opt:
                utilities_sheet.cell(row=5, column=i+2).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Researcher expected utilities
        utilities_sheet['A7'] = "Researcher's expected utilities"
        utilities_sheet['A7'].font = Font(bold=True)
        utilities_sheet['A8'] = "Strategy"
        for i, s in enumerate(strategies):
            utilities_sheet.cell(row=8, column=i+2).value = f"{s[0]},{s[1]}"
            utilities_sheet.cell(row=9, column=i+2).value = expected_utilities_R[s]
            # Highlight optimal strategy
            if s == s_R_opt:
                utilities_sheet.cell(row=9, column=i+2).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Belief Update Sheet
        belief_update_sheet = workbook.create_sheet("Belief Update")
        belief_update_sheet['A1'] = "Belief Update"
        belief_update_sheet['A1'].font = Font(bold=True, size=14)
        
        # Organization's belief update
        belief_update_sheet['A3'] = "Organization's belief update"
        belief_update_sheet['A3'].font = Font(bold=True)
        belief_update_sheet['A4'] = "Strategy"
        belief_update_sheet['B4'] = "Likelihood"
        belief_update_sheet['C4'] = "Unnormalized Posterior"
        belief_update_sheet['D4'] = "Normalization Factor"
        belief_update_sheet['E4'] = "Updated Belief"
        
        for i, s in enumerate(strategies):
            belief_update_sheet.cell(row=i+5, column=1).value = f"{s[0]},{s[1]}"
            belief_update_sheet.cell(row=i+5, column=2).value = likelihood_O[s]
            belief_update_sheet.cell(row=i+5, column=3).value = unnormalized_posterior_O[s]
            belief_update_sheet.cell(row=i+5, column=4).value = normalization_factor_O
            belief_update_sheet.cell(row=i+5, column=5).value = updated_beliefs_O[s]
            
            # Color coding for updated beliefs
            if s == s_R_opt:  # If this is the strategy chosen by the researcher
                belief_update_sheet.cell(row=i+5, column=5).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
            elif updated_beliefs_O[s] > beliefs_O[s]:  # If belief increased
                belief_update_sheet.cell(row=i+5, column=5).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Light yellow
            elif updated_beliefs_O[s] < beliefs_O[s]:  # If belief decreased
                belief_update_sheet.cell(row=i+5, column=5).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red
        
        # Add sum check
        belief_update_sheet.cell(row=len(strategies)+5, column=4).value = "Sum check:"
        belief_update_sheet.cell(row=len(strategies)+5, column=5).value = sum(updated_beliefs_O.values())
        
        # Researcher's belief update
        belief_update_sheet['A13'] = "Researcher's belief update"
        belief_update_sheet['A13'].font = Font(bold=True)
        belief_update_sheet['A14'] = "Strategy"
        belief_update_sheet['B14'] = "Likelihood"
        belief_update_sheet['C14'] = "Unnormalized Posterior"
        belief_update_sheet['D14'] = "Normalization Factor"
        belief_update_sheet['E14'] = "Updated Belief"
        
        for i, s in enumerate(strategies):
            belief_update_sheet.cell(row=i+15, column=1).value = f"{s[0]},{s[1]}"
            belief_update_sheet.cell(row=i+15, column=2).value = likelihood_R[s]
            belief_update_sheet.cell(row=i+15, column=3).value = unnormalized_posterior_R[s]
            belief_update_sheet.cell(row=i+15, column=4).value = normalization_factor_R
            belief_update_sheet.cell(row=i+15, column=5).value = updated_beliefs_R[s]
            
            # Color coding for updated beliefs
            if s == s_O_opt:  # If this is the strategy chosen by the organization
                belief_update_sheet.cell(row=i+15, column=5).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
            elif updated_beliefs_R[s] > beliefs_R[s]:  # If belief increased
                belief_update_sheet.cell(row=i+15, column=5).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Light yellow
            elif updated_beliefs_R[s] < beliefs_R[s]:  # If belief decreased
                belief_update_sheet.cell(row=i+15, column=5).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red
        
        # Add sum check
        belief_update_sheet.cell(row=len(strategies)+15, column=4).value = "Sum check:"
        belief_update_sheet.cell(row=len(strategies)+15, column=5).value = sum(updated_beliefs_R.values())
        
        # Exigence Update Sheet
        exigence_sheet = workbook.create_sheet("Exigence Update")
        exigence_sheet['A1'] = "Exigence Update"
        exigence_sheet['A1'].font = Font(bold=True, size=14)
        
        exigence_sheet['A3'] = "Observed Impact O"
        exigence_sheet['B3'] = observed_impact_O
        exigence_sheet['C3'] = "Expected Impact O"
        exigence_sheet['D3'] = expected_impacts_O[s_O_opt]
        
        exigence_sheet['A4'] = "Observed Impact R"
        exigence_sheet['B4'] = observed_impact_R
        exigence_sheet['C4'] = "Expected Impact R"
        exigence_sheet['D4'] = expected_impacts_R[s_R_opt]
        
        exigence_sheet['A5'] = "Total Observed Impact"
        exigence_sheet['B5'] = observed_total_impact
        exigence_sheet['C5'] = "Total Expected Impact"
        exigence_sheet['D5'] = expected_impacts_O[s_O_opt] + expected_impacts_R[s_R_opt]
        
        exigence_sheet['A7'] = "Current Exigence (X_t)"
        exigence_sheet['B7'] = params.X_init
        
        exigence_sheet['A8'] = "Updated Exigence (X_t+1)"
        exigence_sheet['B8'] = X_next
        
        exigence_sheet['A9'] = "Exigence Change"
        exigence_sheet['B9'] = exigence_change
        # Color code exigence change
        if exigence_change < 0:
            exigence_sheet['B9'].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red
            exigence_sheet['C9'] = "Feedback has dampening effect"
        else:
            exigence_sheet['B9'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
            exigence_sheet['C9'] = "Feedback has amplifying effect"
        
        # Save the workbook
        workbook.save(excel_filename)
        print(f"Saved calculations to {excel_filename}")
    
    return {
        'step': t,
        'X_t': params.X_init,
        'X_next': X_next,
        'strategy_O': s_O_opt,
        'strategy_R': s_R_opt,
        'observed_impact_O': observed_impact_O,
        'observed_impact_R': observed_impact_R,
        'observed_total_impact': observed_total_impact,
        'expected_utilities_O': expected_utilities_O,
        'expected_utilities_R': expected_utilities_R,
        'recalculated_utilities_O': recalculated_utilities_O,  # Add the recalculated utilities
        'recalculated_utilities_R': recalculated_utilities_R,  # Add the recalculated utilities
        'normalized_impact_O': normalized_impact_O,
        'normalized_impact_R': normalized_impact_R,
        'beliefs_O': beliefs_O,
        'beliefs_R': beliefs_R,
        'likelihood_O': likelihood_O,
        'likelihood_R': likelihood_R,
        'unnormalized_posterior_O': unnormalized_posterior_O,
        'unnormalized_posterior_R': unnormalized_posterior_R,
        'normalization_factor_O': normalization_factor_O,
        'normalization_factor_R': normalization_factor_R,
        'updated_beliefs_O': updated_beliefs_O,
        'updated_beliefs_R': updated_beliefs_R,
        'exigence_change': exigence_change
    }

# Simulate multiple steps of the model
def simulate_multi_step(params, steps=10, export_excel=False):
    """
    Simulates multiple steps of the model, updating beliefs and exigence at each step
    """
    results = []
    
    # Initial step
    current_beliefs_O = None
    current_beliefs_R = None
    current_X = params.X_init
    
    for t in range(steps):
        # Set current exigence
        params.X_init = current_X
        
        # Store previous beliefs for Excel output (if not initial step)
        prev_beliefs_O = current_beliefs_O if t > 0 else None
        prev_beliefs_R = current_beliefs_R if t > 0 else None
        
        # Calculate current step
        step_result = calculate_game_step(params, beliefs_O=current_beliefs_O, beliefs_R=current_beliefs_R, 
                                         t=t, export_excel=export_excel,
                                         prev_beliefs_O=prev_beliefs_O, prev_beliefs_R=prev_beliefs_R)
        
        # Record results
        results.append({
            'step': t,
            'X_t': current_X,
            'X_next': step_result['X_next'],
            'strategy_O': step_result['strategy_O'],
            'strategy_R': step_result['strategy_R'],
            'observed_impact': step_result['observed_total_impact'],
            'utility_O': step_result['expected_utilities_O'][step_result['strategy_O']],
            'utility_R': step_result['expected_utilities_R'][step_result['strategy_R']]
        })
        
        # Update for next step
        current_X = step_result['X_next']
        current_beliefs_O = step_result['updated_beliefs_O']
        current_beliefs_R = step_result['updated_beliefs_R']
    
    return pd.DataFrame(results)

# ============ VISUALIZATION FUNCTIONS ============

# Figure 1: Comparison of the two scenarios (Citizen Science vs. Generative AI)
def create_scenarios_comparison(save_pdf=True):
    """
    Create a visualization comparing the two scenarios side by side
    """
    # Create parameters for both scenarios
    params_ai = ModelParams('gen_ai')
    params_cs = ModelParams('citizen_science')
    
    # Calculate optimal strategies and impact for both scenarios
    result_ai = calculate_game_step(params_ai, export_excel=True)
    result_cs = calculate_game_step(params_cs, export_excel=True)
    
    # Get impact values for both scenarios
    impact_O_ai, impact_R_ai = get_impact_values('gen_ai')
    impact_O_cs, impact_R_cs = get_impact_values('citizen_science')
    
    # Create figure
    fig = plt.figure(figsize=(15, 8), constrained_layout=True)
    gs = gridspec.GridSpec(2, 4, width_ratios=[1, 1, 1, 1], figure=fig)
    
    # Define a custom colormap that goes from light to dark blue
    cmap = LinearSegmentedColormap.from_list('blue_gradient', ['#FFFFFF', '#0343DF'])
    
    # Panel 1: Generative AI - Organization Impact
    ax1 = plt.subplot(gs[0, 0])
    strategies = get_strategies()
    impact_values_O_ai = np.array([impact_O_ai[s] for s in strategies]).reshape(2, 3)
    im1 = ax1.imshow(impact_values_O_ai, cmap=cmap, vmin=0, vmax=1)
    ax1.set_title('AI: Organization Impact')
    ax1.set_xticks(np.arange(3))
    ax1.set_yticks(np.arange(2))
    ax1.set_xticklabels(['k', 'd', 'p'])
    ax1.set_yticklabels(['a', 'e'])
    ax1.set_xlabel('Mode')
    ax1.set_ylabel('Role')
    
    # Add text annotations
    for i in range(2):
        for j in range(3):
            s = (('a', 'e')[i], ('k', 'd', 'p')[j])
            ax1.text(j, i, f"{impact_O_ai[s]:.1f}", ha="center", va="center", 
                    color="white" if impact_O_ai[s] > 0.5 else "black")
    
    # Panel 2: Generative AI - Researcher Impact
    ax2 = plt.subplot(gs[0, 1])
    impact_values_R_ai = np.array([impact_R_ai[s] for s in strategies]).reshape(2, 3)
    im2 = ax2.imshow(impact_values_R_ai, cmap=cmap, vmin=0, vmax=1)
    ax2.set_title('AI: Researcher Impact')
    ax2.set_xticks(np.arange(3))
    ax2.set_yticks(np.arange(2))
    ax2.set_xticklabels(['k', 'd', 'p'])
    ax2.set_yticklabels(['a', 'e'])
    ax2.set_xlabel('Mode')
    ax2.set_ylabel('Role')
    
    # Add text annotations
    for i in range(2):
        for j in range(3):
            s = (('a', 'e')[i], ('k', 'd', 'p')[j])
            ax2.text(j, i, f"{impact_R_ai[s]:.1f}", ha="center", va="center", 
                    color="white" if impact_R_ai[s] > 0.5 else "black")
    
    # Panel 3: Citizen Science - Organization Impact
    ax3 = plt.subplot(gs[0, 2])
    impact_values_O_cs = np.array([impact_O_cs[s] for s in strategies]).reshape(2, 3)
    im3 = ax3.imshow(impact_values_O_cs, cmap=cmap, vmin=0, vmax=1)
    ax3.set_title('CS: Organization Impact')
    ax3.set_xticks(np.arange(3))
    ax3.set_yticks(np.arange(2))
    ax3.set_xticklabels(['k', 'd', 'p'])
    ax3.set_yticklabels(['a', 'e'])
    ax3.set_xlabel('Mode')
    ax3.set_ylabel('Role')
    
    # Add text annotations
    for i in range(2):
        for j in range(3):
            s = (('a', 'e')[i], ('k', 'd', 'p')[j])
            ax3.text(j, i, f"{impact_O_cs[s]:.1f}", ha="center", va="center", 
                    color="white" if impact_O_cs[s] > 0.5 else "black")
    
    # Panel 4: Citizen Science - Researcher Impact
    ax4 = plt.subplot(gs[0, 3])
    impact_values_R_cs = np.array([impact_R_cs[s] for s in strategies]).reshape(2, 3)
    im4 = ax4.imshow(impact_values_R_cs, cmap=cmap, vmin=0, vmax=1)
    ax4.set_title('CS: Researcher Impact')
    ax4.set_xticks(np.arange(3))
    ax4.set_yticks(np.arange(2))
    ax4.set_xticklabels(['k', 'd', 'p'])
    ax4.set_yticklabels(['a', 'e'])
    ax4.set_xlabel('Mode')
    ax4.set_ylabel('Role')
    
    # Add text annotations
    for i in range(2):
        for j in range(3):
            s = (('a', 'e')[i], ('k', 'd', 'p')[j])
            ax4.text(j, i, f"{impact_R_cs[s]:.1f}", ha="center", va="center", 
                    color="white" if impact_R_cs[s] > 0.5 else "black")
    
    # Add a colorbar
    cbar_ax = fig.add_axes([0.92, 0.55, 0.02, 0.3])
    fig.colorbar(im1, cax=cbar_ax, label='Impact')
    
    # Panel 5 & 6: Exigence Dynamics for both scenarios
    ax5 = plt.subplot(gs[1, 0:2])
    
    # Simulate 15 steps for both scenarios
    df_ai = simulate_multi_step(params_ai, steps=15, export_excel=True)
    df_cs = simulate_multi_step(params_cs, steps=15, export_excel=True)
    
    # Plot exigence evolution
    ax5.plot(df_ai['step'], df_ai['X_t'], 'b-o', label='AI Exigence')
    ax5.plot(df_cs['step'], df_cs['X_t'], 'g-o', label='CS Exigence')
    
    # Plot impact
    ax5.plot(df_ai['step'], df_ai['observed_impact'], 'b--', label='AI Impact')
    ax5.plot(df_cs['step'], df_cs['observed_impact'], 'g--', label='CS Impact')
    
    # Threshold lines
    ax5.axhline(y=params_ai.theta, color='b', linestyle=':', alpha=0.5, label='AI Threshold')
    ax5.axhline(y=params_cs.theta, color='g', linestyle=':', alpha=0.5, label='CS Threshold')
    
    ax5.set_title('Exigence and Impact Evolution')
    ax5.set_xlabel('Time Steps')
    ax5.set_ylabel('Value')
    ax5.set_ylim(0, 1)
    ax5.legend()
    ax5.grid(True, linestyle='--', alpha=0.7)
    
    # Panel 7: Strategy choices over time for AI
    ax6 = plt.subplot(gs[1, 2])
    
    # Create a mapping of strategies to integers
    strategies = get_strategies()
    strat_to_num = {s: i for i, s in enumerate(strategies)}
    
    # Plot strategy choices
    ax6.plot(df_ai['step'], [strat_to_num[s] for s in df_ai['strategy_O']], 'bs-', label='Org Strategy')
    ax6.plot(df_ai['step'], [strat_to_num[s] for s in df_ai['strategy_R']], 'rs-', label='Res Strategy')
    
    # Set y-ticks to strategy labels
    ax6.set_yticks(range(len(strategies)))
    ax6.set_yticklabels([f"{s[0]},{s[1]}" for s in strategies])
    
    ax6.set_title('AI Scenario: Strategy Evolution')
    ax6.set_xlabel('Time Steps')
    ax6.set_ylabel('Strategy')
    ax6.legend()
    ax6.grid(True, linestyle='--', alpha=0.7)
    
    # Panel 8: Strategy choices over time for CS
    ax7 = plt.subplot(gs[1, 3])
    
    # Plot strategy choices
    ax7.plot(df_cs['step'], [strat_to_num[s] for s in df_cs['strategy_O']], 'bs-', label='Org Strategy')
    ax7.plot(df_cs['step'], [strat_to_num[s] for s in df_cs['strategy_R']], 'rs-', label='Res Strategy')
    
    # Set y-ticks to strategy labels
    ax7.set_yticks(range(len(strategies)))
    ax7.set_yticklabels([f"{s[0]},{s[1]}" for s in strategies])
    
    ax7.set_title('CS Scenario: Strategy Evolution')
    ax7.set_xlabel('Time Steps')
    ax7.set_ylabel('Strategy')
    ax7.legend()
    ax7.grid(True, linestyle='--', alpha=0.7)
    
    #plt.tight_layout()
    
    # Add scenario summaries
    # AI summary
    fig.text(0.25, 0.48, f"AI Optimal Strategies: Org={result_ai['strategy_O']}, Res={result_ai['strategy_R']}", 
             ha='center', fontsize=10, bbox=dict(facecolor='white', alpha=0.5))
    
    # CS summary
    fig.text(0.75, 0.48, f"CS Optimal Strategies: Org={result_cs['strategy_O']}, Res={result_cs['strategy_R']}", 
             ha='center', fontsize=10, bbox=dict(facecolor='white', alpha=0.5))
    
    # Save as PDF
    if save_pdf:
        plt.savefig('simulations/figure1_scenarios_comparison.pdf', format='pdf', bbox_inches='tight')
        print("Saved scenarios comparison as 'simulations/figure1_scenarios_comparison.pdf'")
    
    return fig

# Figure 2: Model development and equilibrium analysis
# New function to create parameter comparison between both scenarios
def create_parameter_comparison(save_pdf=True):
    """
    Create a visualization comparing parameter adjustments between both scenarios
    """
    # Create figure with constrained layout
    fig = plt.figure(figsize=(15, 10), constrained_layout=True)
    gs = gridspec.GridSpec(2, 2, figure=fig)
    
    # Create parameter grid
    alpha_values = np.linspace(0.1, 0.9, 9)
    strategies = get_strategies()
    
    # Panel A: Organization Strategy (AI Scenario)
    params_ai = ModelParams('gen_ai')
    strategy_matrix_O_ai = np.empty((len(alpha_values), len(alpha_values)), dtype=object)
    numeric_matrix_O_ai = np.zeros((len(alpha_values), len(alpha_values)))
    
    # Calculate optimal strategies for each alpha combination (AI)
    for i, alpha_O in enumerate(alpha_values):
        for j, alpha_R in enumerate(alpha_values):
            params_ai.alpha_O = alpha_O
            params_ai.alpha_R = alpha_R
            
            result = calculate_game_step(params_ai)
            strategy_matrix_O_ai[i, j] = result['strategy_O']
    
    # Convert strategies to numeric codes for visualization
    strategy_to_num = {s: i for i, s in enumerate(strategies)}
    for i in range(len(alpha_values)):
        for j in range(len(alpha_values)):
            numeric_matrix_O_ai[i, j] = strategy_to_num[strategy_matrix_O_ai[i, j]]
    
    # Panel B: Organization Strategy (CS Scenario)
    params_cs = ModelParams('citizen_science')
    strategy_matrix_O_cs = np.empty((len(alpha_values), len(alpha_values)), dtype=object)
    numeric_matrix_O_cs = np.zeros((len(alpha_values), len(alpha_values)))
    
    # Calculate optimal strategies for each alpha combination (CS)
    for i, alpha_O in enumerate(alpha_values):
        for j, alpha_R in enumerate(alpha_values):
            params_cs.alpha_O = alpha_O
            params_cs.alpha_R = alpha_R
            
            result = calculate_game_step(params_cs)
            strategy_matrix_O_cs[i, j] = result['strategy_O']
    
    # Convert strategies to numeric codes for visualization
    for i in range(len(alpha_values)):
        for j in range(len(alpha_values)):
            numeric_matrix_O_cs[i, j] = strategy_to_num[strategy_matrix_O_cs[i, j]]
    
    # Panel C: Researcher Strategy (AI Scenario)
    params_ai = ModelParams('gen_ai')
    strategy_matrix_R_ai = np.empty((len(alpha_values), len(alpha_values)), dtype=object)
    numeric_matrix_R_ai = np.zeros((len(alpha_values), len(alpha_values)))
    
    # Calculate optimal strategies for each alpha combination (AI)
    for i, alpha_O in enumerate(alpha_values):
        for j, alpha_R in enumerate(alpha_values):
            params_ai.alpha_O = alpha_O
            params_ai.alpha_R = alpha_R
            
            result = calculate_game_step(params_ai)
            strategy_matrix_R_ai[i, j] = result['strategy_R']
    
    # Convert strategies to numeric codes for visualization
    for i in range(len(alpha_values)):
        for j in range(len(alpha_values)):
            numeric_matrix_R_ai[i, j] = strategy_to_num[strategy_matrix_R_ai[i, j]]
    
    # Panel D: Researcher Strategy (CS Scenario)
    params_cs = ModelParams('citizen_science')
    strategy_matrix_R_cs = np.empty((len(alpha_values), len(alpha_values)), dtype=object)
    numeric_matrix_R_cs = np.zeros((len(alpha_values), len(alpha_values)))
    
    # Calculate optimal strategies for each alpha combination (CS)
    for i, alpha_O in enumerate(alpha_values):
        for j, alpha_R in enumerate(alpha_values):
            params_cs.alpha_O = alpha_O
            params_cs.alpha_R = alpha_R
            
            result = calculate_game_step(params_cs)
            strategy_matrix_R_cs[i, j] = result['strategy_R']
    
    # Convert strategies to numeric codes for visualization
    for i in range(len(alpha_values)):
        for j in range(len(alpha_values)):
            numeric_matrix_R_cs[i, j] = strategy_to_num[strategy_matrix_R_cs[i, j]]
    
    # Plot the four heatmaps
    ax1 = plt.subplot(gs[0, 0])
    im1 = ax1.imshow(numeric_matrix_O_ai, cmap='viridis', origin='lower')
    ax1.set_title('A: Organization Strategy - Generative AI')
    ax1.set_xlabel('Researcher alpha (α_R)')
    ax1.set_ylabel('Organization alpha (α_O)')
    ax1.set_xticks(np.arange(len(alpha_values)))
    ax1.set_yticks(np.arange(len(alpha_values)))
    ax1.set_xticklabels([f"{alpha:.1f}" for alpha in alpha_values])
    ax1.set_yticklabels([f"{alpha:.1f}" for alpha in alpha_values])
    
    # Add strategy labels
    for i in range(len(alpha_values)):
        for j in range(len(alpha_values)):
            ax1.text(j, i, f"{strategy_matrix_O_ai[i, j][0]},{strategy_matrix_O_ai[i, j][1]}", 
                    ha="center", va="center", color="w", fontsize=8)
    
    ax2 = plt.subplot(gs[0, 1])
    im2 = ax2.imshow(numeric_matrix_O_cs, cmap='viridis', origin='lower')
    ax2.set_title('B: Organization Strategy - Citizen Science')
    ax2.set_xlabel('Researcher alpha (α_R)')
    ax2.set_ylabel('Organization alpha (α_O)')
    ax2.set_xticks(np.arange(len(alpha_values)))
    ax2.set_yticks(np.arange(len(alpha_values)))
    ax2.set_xticklabels([f"{alpha:.1f}" for alpha in alpha_values])
    ax2.set_yticklabels([f"{alpha:.1f}" for alpha in alpha_values])
    
    # Add strategy labels
    for i in range(len(alpha_values)):
        for j in range(len(alpha_values)):
            ax2.text(j, i, f"{strategy_matrix_O_cs[i, j][0]},{strategy_matrix_O_cs[i, j][1]}", 
                    ha="center", va="center", color="w", fontsize=8)
    
    ax3 = plt.subplot(gs[1, 0])
    im3 = ax3.imshow(numeric_matrix_R_ai, cmap='viridis', origin='lower')
    ax3.set_title('C: Researcher Strategy - Generative AI')
    ax3.set_xlabel('Researcher alpha (α_R)')
    ax3.set_ylabel('Organization alpha (α_O)')
    ax3.set_xticks(np.arange(len(alpha_values)))
    ax3.set_yticks(np.arange(len(alpha_values)))
    ax3.set_xticklabels([f"{alpha:.1f}" for alpha in alpha_values])
    ax3.set_yticklabels([f"{alpha:.1f}" for alpha in alpha_values])
    
    # Add strategy labels
    for i in range(len(alpha_values)):
        for j in range(len(alpha_values)):
            ax3.text(j, i, f"{strategy_matrix_R_ai[i, j][0]},{strategy_matrix_R_ai[i, j][1]}", 
                    ha="center", va="center", color="w", fontsize=8)
    
    ax4 = plt.subplot(gs[1, 1])
    im4 = ax4.imshow(numeric_matrix_R_cs, cmap='viridis', origin='lower')
    ax4.set_title('D: Researcher Strategy - Citizen Science')
    ax4.set_xlabel('Researcher alpha (α_R)')
    ax4.set_ylabel('Organization alpha (α_O)')
    ax4.set_xticks(np.arange(len(alpha_values)))
    ax4.set_yticks(np.arange(len(alpha_values)))
    ax4.set_xticklabels([f"{alpha:.1f}" for alpha in alpha_values])
    ax4.set_yticklabels([f"{alpha:.1f}" for alpha in alpha_values])
    
    # Add strategy labels
    for i in range(len(alpha_values)):
        for j in range(len(alpha_values)):
            ax4.text(j, i, f"{strategy_matrix_R_cs[i, j][0]},{strategy_matrix_R_cs[i, j][1]}", 
                    ha="center", va="center", color="w", fontsize=8)
    
    # Add colorbar
    cbar_ax = fig.add_axes([0.92, 0.15, 0.02, 0.7])
    fig.colorbar(im1, cax=cbar_ax, label='Strategy Index')
    
    # Save as PDF
    if save_pdf:
        plt.savefig('simulations/figure2_parameter_comparison.pdf', format='pdf')
        print("Saved parameter comparison as 'simulations/figure2_parameter_comparison.pdf'")
    
    return fig

# New function to create exigence dynamics comparison
def create_exigence_dynamics_comparison(save_pdf=True):
    """
    Create visualization comparing exigence dynamics for different parameters between both scenarios
    """
    # Create figure with constrained layout
    fig = plt.figure(figsize=(15, 6), constrained_layout=True)
    gs = gridspec.GridSpec(1, 2, figure=fig)
    
    # Panel A: Exigence dynamics for Generative AI
    ax1 = plt.subplot(gs[0, 0])
    
    # Define different parameter sets
    param_sets = [
        {'beta': 0.2, 'theta': 0.3, 'label': 'Low β, Low θ', 'color': 'b'},
        {'beta': 0.5, 'theta': 0.5, 'label': 'Medium β, Medium θ', 'color': 'g'},
        {'beta': 0.8, 'theta': 0.7, 'label': 'High β, High θ', 'color': 'r'}
    ]
    
    # Simulate for Generative AI
    for param_set in param_sets:
        params_ai = ModelParams('gen_ai')
        params_ai.beta = param_set['beta']
        params_ai.theta = param_set['theta']
        
        df_ai = simulate_multi_step(params_ai, steps=20)
        
        ax1.plot(df_ai['step'], df_ai['X_t'], f"{param_set['color']}o-", 
                label=param_set['label'])
        
        # Add threshold lines
        ax1.axhline(y=param_set['theta'], color=param_set['color'], 
                  linestyle=':', alpha=0.5)
    
    ax1.set_title('A: Exigence Dynamics - Generative AI')
    ax1.set_xlabel('Time Steps')
    ax1.set_ylabel('Exigence')
    ax1.set_ylim(0, 1)
    ax1.legend()
    ax1.grid(True, linestyle='--', alpha=0.7)
    
    # Panel B: Exigence dynamics for Citizen Science
    ax2 = plt.subplot(gs[0, 1])
    
    # Simulate for Citizen Science
    for param_set in param_sets:
        params_cs = ModelParams('citizen_science')
        params_cs.beta = param_set['beta']
        params_cs.theta = param_set['theta']
        
        df_cs = simulate_multi_step(params_cs, steps=20)
        
        ax2.plot(df_cs['step'], df_cs['X_t'], f"{param_set['color']}o-", 
                label=param_set['label'])
        
        # Add threshold lines
        ax2.axhline(y=param_set['theta'], color=param_set['color'], 
                  linestyle=':', alpha=0.5)
    
    ax2.set_title('B: Exigence Dynamics - Citizen Science')
    ax2.set_xlabel('Time Steps')
    ax2.set_ylabel('Exigence')
    ax2.set_ylim(0, 1)
    ax2.legend()
    ax2.grid(True, linestyle='--', alpha=0.7)
    
    # Save as PDF
    if save_pdf:
        plt.savefig('simulations/figure3_exigence_dynamics_comparison.pdf', format='pdf')
        print("Saved exigence dynamics comparison as 'simulations/figure3_exigence_dynamics_comparison.pdf'")
    
    return fig

# Function to create detailed scenario analysis
def create_detailed_scenario_analysis(scenario='gen_ai', save_pdf=True):
    """
    Create a detailed analysis visualization for a specific scenario
    
    Parameters:
    - scenario: 'gen_ai' or 'citizen_science'
    """
    # Set up parameters
    params = ModelParams(scenario)
    
    # Scenario name for plot titles
    scenario_name = "Generative AI" if scenario == 'gen_ai' else "Citizen Science"
    scenario_abbr = 'AI' if scenario == 'gen_ai' else 'CS'
    
    # Create figure with constrained layout
    fig = plt.figure(figsize=(15, 10), constrained_layout=True)
    gs = gridspec.GridSpec(2, 2, figure=fig)
    
    # Panel A: Time evolution of utilities and impact
    ax1 = plt.subplot(gs[0, 0])
    
    # Simulate 15 steps
    df = simulate_multi_step(params, steps=15)
    
    # Plot utilities and impact
    ax1.plot(df['step'], df['utility_O'], 'bo-', label='Organization Utility')
    ax1.plot(df['step'], df['utility_R'], 'ro-', label='Researcher Utility')
    ax1.plot(df['step'], df['observed_impact'], 'go-', label='Combined Impact')
    ax1.axhline(y=params.theta, color='k', linestyle='--', alpha=0.5, label='Threshold (θ)')
    
    ax1.set_title(f'A: Evolution of Utilities and Impact - {scenario_name}')
    ax1.set_xlabel('Time Steps')
    ax1.set_ylabel('Value')
    ax1.legend()
    ax1.grid(True, linestyle='--', alpha=0.7)
    
    # Panel B: Strategy choices over time
    ax2 = plt.subplot(gs[0, 1])
    
    # Create a mapping of strategies to integers
    strategies = get_strategies()
    strat_to_num = {s: i for i, s in enumerate(strategies)}
    
    # Plot strategy choices
    ax2.plot(df['step'], [strat_to_num[s] for s in df['strategy_O']], 'bs-', label='Org Strategy')
    ax2.plot(df['step'], [strat_to_num[s] for s in df['strategy_R']], 'rs-', label='Res Strategy')
    
    # Set y-ticks to strategy labels
    ax2.set_yticks(range(len(strategies)))
    ax2.set_yticklabels([f"{s[0]},{s[1]}" for s in strategies])
    
    ax2.set_title(f'B: Strategy Evolution - {scenario_name}')
    ax2.set_xlabel('Time Steps')
    ax2.set_ylabel('Strategy')
    ax2.legend()
    ax2.grid(True, linestyle='--', alpha=0.7)
    
    # Panel C: Equilibrium analysis for different theta values
    ax3 = plt.subplot(gs[1, 0])
    
    # Analyze equilibrium for different theta values
    theta_values = np.linspace(0.1, 0.9, 9)
    equilibrium_exigence = []
    equilibrium_impact = []
    strategy_changes_O = []
    strategy_changes_R = []
    
    for theta in theta_values:
        params_eq = ModelParams(scenario)
        params_eq.theta = theta
        
        # Run a simulation to reach equilibrium
        df_eq = simulate_multi_step(params_eq, steps=30)
        
        # Get the last few values (should be near equilibrium)
        last_exigence = df_eq['X_t'].iloc[-5:].mean()
        last_impact = df_eq['observed_impact'].iloc[-5:].mean()
        last_strategy_O = df_eq['strategy_O'].iloc[-1]
        last_strategy_R = df_eq['strategy_R'].iloc[-1]
        
        equilibrium_exigence.append(last_exigence)
        equilibrium_impact.append(last_impact)
        strategy_changes_O.append(last_strategy_O)
        strategy_changes_R.append(last_strategy_R)
    
    # Plot equilibrium values
    ax3.plot(theta_values, equilibrium_exigence, 'bo-', label='Equilibrium Exigence')
    ax3.plot(theta_values, equilibrium_impact, 'ro-', label='Equilibrium Impact')
    ax3.plot(theta_values, theta_values, 'k--', alpha=0.5, label='Threshold (θ)')
    
    # Annotate strategy regions
    strategy_regions_O = []
    strategy_regions_R = []
    current_strat_O = strategy_changes_O[0]
    current_strat_R = strategy_changes_R[0]
    region_start_O = theta_values[0]
    region_start_R = theta_values[0]
    
    for i, (strat_O, strat_R, theta) in enumerate(zip(strategy_changes_O, strategy_changes_R, theta_values)):
        # Check for organization strategy change
        if strat_O != current_strat_O or i == len(strategy_changes_O) - 1:
            strategy_regions_O.append((region_start_O, theta, current_strat_O))
            region_start_O = theta
            current_strat_O = strat_O
        
        # Check for researcher strategy change
        if strat_R != current_strat_R or i == len(strategy_changes_R) - 1:
            strategy_regions_R.append((region_start_R, theta, current_strat_R))
            region_start_R = theta
            current_strat_R = strat_R
    
    # Annotate strategy regions
    y_pos_O = 0.85
    y_pos_R = 0.75
    for start, end, strat in strategy_regions_O:
        ax3.annotate(f"O:{strat}", xy=((start + end)/2, y_pos_O), 
                   xytext=(0, 0), textcoords='offset points', 
                   ha='center', va='center', fontsize=8,
                   bbox=dict(boxstyle="round,pad=0.3", fc="skyblue", alpha=0.3))
    
    for start, end, strat in strategy_regions_R:
        ax3.annotate(f"R:{strat}", xy=((start + end)/2, y_pos_R), 
                   xytext=(0, 0), textcoords='offset points', 
                   ha='center', va='center', fontsize=8,
                   bbox=dict(boxstyle="round,pad=0.3", fc="lightcoral", alpha=0.3))
    
    ax3.set_title(f'C: Equilibrium Points for Different Threshold Values - {scenario_name}')
    ax3.set_xlabel('Threshold Impact (θ)')
    ax3.set_ylabel('Equilibrium Value')
    ax3.set_ylim(0, 1)
    ax3.legend()
    ax3.grid(True, linestyle='--', alpha=0.7)
    
    # Panel D: Impact distribution for different alpha values
    ax4 = plt.subplot(gs[1, 1])
    
    # Calculate impact for a range of alpha values
    alpha_range = np.linspace(0.1, 0.9, 9)
    impacts = []
    strategies_O = []
    strategies_R = []
    
    for alpha in alpha_range:
        params_impact = ModelParams(scenario)
        params_impact.alpha_O = alpha
        params_impact.alpha_R = alpha
        
        result = calculate_game_step(params_impact)
        
        impacts.append(result['observed_total_impact'])
        strategies_O.append(result['strategy_O'])
        strategies_R.append(result['strategy_R'])
    
    # Create a bar chart
    ax4.bar(alpha_range, impacts, color='skyblue', alpha=0.7)
    
    # Add threshold line
    ax4.axhline(y=params.theta, color='r', linestyle='--', 
              label=f'Threshold (θ = {params.theta})')
    
    # Add strategy annotations
    for i, (alpha, impact, strat_O, strat_R) in enumerate(zip(alpha_range, impacts, strategies_O, strategies_R)):
        ax4.annotate(f"O:{strat_O}\nR:{strat_R}", 
                   xy=(alpha, impact), 
                   xytext=(0, 10 if i % 2 == 0 else -35), 
                   textcoords='offset points',
                   ha='center', fontsize=8,
                   bbox=dict(boxstyle="round,pad=0.2", fc="white", alpha=0.7))
    
    ax4.set_title(f'D: Impact and Strategies for Different Alpha Values - {scenario_name}')
    ax4.set_xlabel('Alpha (same for both actors)')
    ax4.set_ylabel('Combined Impact')
    ax4.set_ylim(0, 1)
    ax4.legend()
    ax4.grid(True, linestyle='--', alpha=0.7)
    
    # Save as PDF
    if save_pdf:
        plt.savefig(f'simulations/figure_detailed_{scenario_abbr}_analysis.pdf', format='pdf')
        print(f"Saved detailed {scenario_name} analysis as 'simulations/figure_detailed_{scenario_abbr}_analysis.pdf'")
    
    return fig

# Updated function to create all visualizations
def create_all_visualizations(save_pdf=True):
    """
    Create all the reorganized visualizations for the paper
    """
    print("Generating reorganized visualizations for the paper...")
    
    # Create Figure 1: Comparison of the two scenarios
    fig1 = create_scenarios_comparison(save_pdf=save_pdf)
    
    # Create Figure 2: Parameter comparison between both scenarios
    fig2 = create_parameter_comparison(save_pdf=save_pdf)
    
    # Create Figure 3: Exigence dynamics comparison
    fig3 = create_exigence_dynamics_comparison(save_pdf=save_pdf)
    
    # Create detailed scenario analyses
    fig_ai = create_detailed_scenario_analysis(scenario='gen_ai', save_pdf=save_pdf)
    fig_cs = create_detailed_scenario_analysis(scenario='citizen_science', save_pdf=save_pdf)
    
    print("All visualizations complete. PDF files saved in the 'simulations/' directory")
    
    return {
        'fig1': fig1, 
        'fig2': fig2,
        'fig3': fig3,
        'fig_ai': fig_ai,
        'fig_cs': fig_cs
    }


# Run directly to create all visualizations
if __name__ == "__main__":
    # Check if simulations directory exists
    if not os.path.exists('simulations'):
        os.makedirs('simulations')
    
    # Set the current working directory
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    os.chdir(SCRIPT_DIR)
    print("Current working directory:", os.getcwd())
    
    create_all_visualizations(save_pdf=True)